import cv2
import numpy as np
import os
from tkinter import *
from tkinter import messagebox
from tkinter import simpledialog  
import pandas as pd
from openpyxl import load_workbook
import openpyxl



# Constants
DATA_PATH = 'data'  # Directory to store training data
MODEL_FILENAME = 'face_model.xml'  
FONT = cv2.FONT_HERSHEY_SIMPLEX
FONT_SCALE = 1
FONT_THICKNESS = 2

class FaceDetector:
    def __init__(self):
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.labels_to_names = {}
    def detect_faces(self, img):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
        return faces

    def capture_training_data(self):
        # Capture images for training
        name = simpledialog.askstring("Input", "Enter your name:")
        cap = cv2.VideoCapture(0)
        count = 0
        while True:
            ret, frame = cap.read()
            faces = self.detect_faces(frame)
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
                if not os.path.exists(DATA_PATH):
                    os.makedirs(DATA_PATH)
                
                label = hash(name) % 10000  # Generate label based on name hash
                self.labels_to_names[label] = name
                cv2.imwrite(f'{DATA_PATH}/{name}_{count}.jpg', frame[y:y+h, x:x+w])
                count += 1
            cv2.imshow('Capturing Training Data', frame)
            if cv2.waitKey(20) & 0xFF == ord('q') or count >= 100:
                break
        cap.release()
        cv2.destroyAllWindows()
        messagebox.showinfo("Information", "Training Data Captured Successfully.")

    def train_model(self):
        # Load training data
        data = []
        labels = []
        for filename in os.listdir(DATA_PATH):
            if filename.endswith('.jpg'):
                name = filename.split('_')[0]  # Extract name from filename
                label = hash(name) % 10000  # Generate label based on name hash
                self.labels_to_names[label] = name
                img = cv2.imread(os.path.join(DATA_PATH, filename))
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                data.append(gray)
                labels.append(label)


        face_recognizer = cv2.face.LBPHFaceRecognizer_create()
        face_recognizer.train(data, np.array(labels))

       
        face_recognizer.save(MODEL_FILENAME)
        messagebox.showinfo("Information", "Model Trained Successfully.")

    def test_model(self):
        # Load trained model
        face_recognizer = cv2.face.LBPHFaceRecognizer_create()
        face_recognizer.read(MODEL_FILENAME)
        index=1
        
        cap = cv2.VideoCapture(0)
        while True:
            ret, frame = cap.read()
            faces = self.detect_faces(frame)
            for (x, y, w, h) in faces:
                gray_face = cv2.cvtColor(frame[y:y+h, x:x+w], cv2.COLOR_BGR2GRAY)
                label, confidence = face_recognizer.predict(gray_face)
                if confidence < 100:  
                    person_name = self.labels_to_names.get(label, "Unknown")
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                    cv2.putText(frame, person_name, (x, y-10), FONT, FONT_SCALE, (0, 255, 0), FONT_THICKNESS)

                    wb = openpyxl.load_workbook("Attendance.xlsx")
                    ws = wb.active
                    existing_data = set(ws[f'A{i}'].value for i in range(1, ws.max_row + 1))

                    if person_name not in existing_data:
                        index = ws.max_row + 1
                        ws[f'A{index}'] = person_name
                        wb.save("Attendance.xlsx")
                else:
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                    cv2.putText(frame, 'Unknown', (x, y-10), FONT, FONT_SCALE, (0, 0, 255), FONT_THICKNESS)
            cv2.imshow('Face Recognition', frame)
            if cv2.waitKey(20) & 0xFF == ord('q'):
                break
        cap.release()
        cv2.destroyAllWindows()
    def clearxl(self):
        wb = openpyxl.load_workbook("Attendance.xlsx")
        ws = wb.active
        for i in range(1, ws.max_row + 1):
            ws[f'A{i}'] = None
        wb.save("Attendance.xlsx")
        messagebox.showinfo("Information", "Attendance Sheet Cleared Successfully.")


class FaceApp:
    def __init__(self, master):
        self.master = master
        master.title("Smart Attendance System")


        self.detector = FaceDetector()

        self.title= Label(master,text="Smart Attendance System",font=("20px"),fg="#191923",bg="#F5F5DC")
        self.title.pack(padx=20,pady=30)

        self.capture_button = Button(master, text="Enroll Student", command=self.detector.capture_training_data,height=2,width=20,bg="#191923",fg="white")
        self.capture_button.pack(pady=20)

        self.train_button = Button(master, text="Train Model", command=self.detector.train_model,height=2,width=20,bg="#191923",fg="white")
        self.train_button.pack(pady=5)

        self.test_button = Button(master, text="Take Attendance", command=self.detector.test_model,height=2,width=20,bg="#191923",fg="white")
        self.test_button.pack(pady=20)
        self.test_button = Button(master, text="Clear Execel Datasheet", command=self.detector.clearxl,height=2,width=20,bg="#191923",fg="white")
        self.test_button.pack(pady=20)


def main():
    root = Tk()
    app = FaceApp(root)
    root.geometry("800x600")
    root.configure(bg="#F5F5DC")
    root.mainloop()


if __name__ == "__main__":
    main()
